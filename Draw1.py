import openpyxl
from openpyxl import Workbook
import openpyxl as openpyxl
from openpyxl.chart import BarChart, Series, Reference
wb = Workbook(write_only=True)
from openpyxl.styles import fonts,colors
from openpyxl.utils import get_column_letter, column_index_from_string
wb=openpyxl.load_workbook('/Users/mac/Desktop/stu_scores _Grade 2.xlsx')
sheet=wb['stu_scores_01']

ob1=openpyxl.chart.Reference(sheet,min_col=3,min_row=34,max_row=34,max_col=7)
ob2=openpyxl.chart.Reference(sheet,min_col=3,min_row=1,max_row=7,max_col=1)

serobj=openpyxl.chart.Series(ob1,title="average score of every subjects",title_from_data=True)

charObj=openpyxl.chart.BarChart()
charObj.title="My Bar Chart"
charObj.append(serobj)
charObj.set_categories(ob2)
sheet.add_chart(charObj,'l2')

ob3=openpyxl.chart.Reference(sheet,min_col=2,min_row=37,max_row=37)
ob4=openpyxl.chart.Reference(sheet,min_col=1,min_row=37,max_row=37)

serobj=openpyxl.chart.Series(ob3,title="average score of every subjects",title_from_data=True)
wb.save('stu_scores _Grade.xlsx')
wb.close()
print("Done")