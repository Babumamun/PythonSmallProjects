import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
wb=openpyxl.load_workbook('/Users/mac/Desktop/stu_scores.xlsx')
sheet = wb['stu_scores']
sheet['H1'] = "Total"
sheet['H1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
sheet['I1'] = "Average"
sheet['I1'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
sheet['J1'] = "Male's Average"
sheet['J1'].fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
sheet['K1'] = "Female's Average"
sheet['K1'].fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
sheet['A34'] = "Highest total score"
sheet['A34'].fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#Total for every students
for i in range(2,sheet.max_row):
    sheet[f'H{i}'].value=f'=SUM(C{i}:G{i})'
#Calculate the average score of each subject
#for i in range(2,sheet.max_row):
#    sheet[f'I{i}']=f'=AVERAGE(C{i}:G{i})'
for i in range(2,7):
    col = get_column_letter(i+1)
    sheet[f'I{i}'].value = f'=AVERAGE({col}2:{col}33)'
#average score of boys in each subject
##for i in range(2,sheet.max_row):
##    sheet[f'J{i}']=f'=AVERAGEIF(B{i},"male",C{i}:G{i})'
for i in range(2,7):
    col = get_column_letter(i+1)
    sheet[f'J{i}'].value= f'=AVERAGEIF(B2:B33,"male",{col}2:{col}33)'

# average score of girls in each subject
#for i in range(2,sheet.max_row):
#    sheet[f'K{i}']=f'=AVERAGEIF(B{i},"female",C{i}:G{i})'
for i in range(2,7):
    col = get_column_letter(i+1)
    sheet[f'K{i}'].value = f'=AVERAGEIF(B2:B33,"female",{col}2:{col}33)'
# student with the highest total scores
sheet['B34']=f"=INDEX(A{2}:A{33},MATCH(MAX(H{2}:H{33}),H{2}:H{33},{0}))"

wb.save('stu_scoresnew.xlsx')
wb.close().value
print("Done")
